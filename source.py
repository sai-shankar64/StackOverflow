import requests
from openpyxl import Workbook,load_workbook

def gtmetrix():
    wb=load_workbook(filename="Format.xlsx")
    worksheets=wb.sheetnames
    no_of_sheets=len(worksheets)
    for sheet in range(0,no_of_sheets):
        ws=wb[worksheets[sheet]]
        for row in ws.rows:
            continue
            for cell in row:
                print(cell.value)
    # data={
    #     "url":"https://www.google.com"
    # }
    # auth=("jayasaishankar@gmail.com","8295189afb8c711f55c454c11aadfb59")
    # response=requests.post("https://gtmetrix.com/api/0.1/test",data=data,auth=auth)
    # print(type(response))
    # print(response.status_code)
    # print(response.json())

if __name__=="__main__":
    gtmetrix()